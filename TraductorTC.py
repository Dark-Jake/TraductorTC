import threading
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from tqdm import tqdm
import time
import os
import sys
import webbrowser
import winsound
import ctypes
import json
import re

ctypes.windll.kernel32.SetConsoleTitleW("Translator by DarkJake v2.0.0")

ascii_art = """
+================================================================================================================+
||   ______   ______     ______     __   __     ______     __         ______     ______   ______     ______     ||
||  /\__  _\ /\  == \   /\  __ \   /\ "-.\ \   /\  ___\   /\ \       /\  __ \   /\__  _\ /\  __ \   /\  == \    ||
||  \/_/\ \/ \ \  __<   \ \  __ \  \ \ \-.  \  \ \___  \  \ \ \____  \ \  __ \  \/_/\ \/ \ \ \/\ \  \ \  __<    ||
||     \ \_\  \ \_\ \_\  \ \_\ \_\  \ \_\ \"\_\  \/\_____\  \ \_____\  \ \_\ \_\    \ \_\  \ \_____\  \ \_\ \_\  ||
||      \/_/   \/_/ /_/   \/_/\/_/   \/_/ \/_/   \/_____/   \/_____/   \/_/\/_/     \/_/   \/_____/   \/_/ /_/  ||
||                                                                                                              ||
||                                                  ______     __  __                                           ||
||                                                 /\  == \   /\ \_\ \                                          ||  
||                                                 \ \  __<   \ \____ \                                         ||  
||                                                  \ \_____\  \/\_____\                                        ||  
||                                                   \/_____/   \/_____/                                        ||  
||                                                                                                              ||  
||                  _____     ______     ______     __  __       __     ______     __  __     ______            ||
||                 /\  __-.  /\  __ \   /\  == \   /\ \/ /      /\ \   /\  __ \   /\ \/ /    /\  ___\           ||
||                 \ \ \/\ \ \ \  __ \  \ \  __<   \ \  _"-.   _\_\ \  \ \  __ \  \ \  _"-.  \ \  __\           ||
||                  \ \____-  \ \_\ \_\  \ \_\ \_\  \ \_\ \_\ /\_____\  \ \_\ \_\  \ \_\ \_\  \ \_____\         ||
||                   \/____/   \/_/\/_/   \/_/ /_/   \/_/\/_/ \/_____/   \/_/\/_/   \/_/\/_/   \/_____/         ||
||                                                                                                              ||  
+================================================================================================================+
"""

def cargar_arte():
    print(ascii_art)

arte_thread = threading.Thread(target=cargar_arte)
arte_thread.start()

arte_thread.join()

def obtener_archivos():
    root = tk.Tk()
    root.withdraw()

    archivo_entrada = filedialog.askopenfilename(title="Selecciona un archivo Excel", filetypes=[("Archivos Excel", "*.xlsx;*.xls")])

    if not archivo_entrada:
        print("No se ha seleccionado ningún archivo.")
        time.sleep(3)
        return None, None

    archivo_salida = filedialog.asksaveasfilename(title="Guardar archivo modificado", defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])

    if not archivo_salida:
        print("No se ha seleccionado una ubicación de guardado.")
        time.sleep(3)
        return None, None

    return archivo_entrada, archivo_salida

def cargar_palabras_desde_json(archivo_json):
    try:
        with open(archivo_json, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            palabras_a_reemplazar = data.get("palabras_a_reemplazar", {})
            return palabras_a_reemplazar
    except Exception as e:
        print(f"Error al cargar el archivo JSON: {str(e)}")
        return {}

def reemplazar_palabras(archivo_entrada, archivo_salida, palabras_a_reemplazar, idioma_origen, idioma_destino):
    wb = load_workbook(archivo_entrada)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        celdas_procesadas = set()

        total_celdas = sheet.max_row * sheet.max_column
        progreso = tqdm(total=total_celdas, desc=f"Procesando hoja: {sheet_name}")

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=1):
            for col_idx, cell in enumerate(row, start=1):
                progreso.update(1) 
                if cell.value is not None and isinstance(cell.value, str):

                    if (row_idx, col_idx) in celdas_procesadas:
                        continue

                    cell_value_lower = cell.value.lower()
                    cell_value_nuevo = cell.value

                    for palabra_original, palabra_nueva in palabras_a_reemplazar.items():
                        palabra_original_lower = palabra_original.lower()
                        if isinstance(palabra_nueva, dict):
                            for categoria, subcategorias in palabra_nueva.items():
                                for subcategoria, valor in subcategorias.items():
                                    if idioma_origen == "inglés":
                                        if subcategoria.lower() in cell_value_lower:
                                            cell_value_nuevo = cell_value_nuevo.replace(subcategoria, valor)
                                    elif idioma_origen == "español":
                                        if valor.lower() in cell_value_lower:
                                            cell_value_nuevo = cell_value_nuevo.replace(valor, subcategoria)
                        else:
                            if idioma_origen == "inglés":
                                cell_value_nuevo = cell_value_nuevo.replace(palabra_original, palabra_nueva)
                            elif idioma_origen == "español":
                                cell_value_nuevo = cell_value_nuevo.replace(palabra_nueva, palabra_original)

                        celdas_procesadas.add((row_idx, col_idx))

                    cell.value = cell_value_nuevo

        progreso.close()

    wb.save(archivo_salida)

def reemplazar_palabras_tc():
    def seleccionar_idioma(idioma):
        nonlocal idioma_origen
        idioma_origen = idioma
        ventana.destroy()

    idioma_origen = None

    ventana = tk.Tk()
    ventana.title("Seleccionar idioma origen")

    etiqueta = tk.Label(ventana, text="¿Desde qué idioma desea traducir?")
    etiqueta.pack()

    boton_espanol = tk.Button(ventana, text="Español", command=lambda: seleccionar_idioma("español"))
    boton_espanol.pack()

    boton_ingles = tk.Button(ventana, text="Inglés", command=lambda: seleccionar_idioma("inglés"))
    boton_ingles.pack()

    ventana.mainloop()

    if not idioma_origen:
        messagebox.showerror("Error", "Por favor, seleccione un idioma.")
        return

    idioma_destino = "español" if idioma_origen == "inglés" else "inglés"
    messagebox.showinfo("Idioma destino", f"Se traducirá desde {idioma_origen} a {idioma_destino}.")

    archivo_entrada, archivo_salida = obtener_archivos()

    if archivo_entrada and archivo_salida:
        archivo_json = 'palabras_a_reemplazar.json'
        palabras_a_reemplazar = cargar_palabras_desde_json(archivo_json)

        reemplazar_palabras(archivo_entrada, archivo_salida, palabras_a_reemplazar, idioma_origen, idioma_destino)

        if hasattr(sys, '_MEIPASS'):
            sound_path = os.path.join(sys._MEIPASS, "TaoAudio_mezcla.wav")
        else:
            sound_path = "TaoAudio_mezcla.wav"

        def reproducir_sonido():
            winsound.PlaySound(sound_path, winsound.SND_NOWAIT)

        sonido_thread = threading.Thread(target=reproducir_sonido)
        sonido_thread.start()

        messagebox.showinfo("Reemplazo completado. El archivo modificado se guardó en:", archivo_salida)
        messagebox.showinfo("Script by", "DarkJake#6238")

        last_run_file = "last_run.txt"
        current_time = time.time()

        if os.path.exists(last_run_file):
            with open(last_run_file) as f:
                last_run = int(round(float(f.read().strip())))
        else:
            last_run = 0

        if current_time - last_run > 600:
            ctypes.windll.kernel32.SetFileAttributesW(last_run_file, 0)
            with open(last_run_file, "w") as f:
                f.write(str(current_time))

            ctypes.windll.kernel32.SetFileAttributesW(last_run_file, 2)

            webbrowser.open("https://www.youtube.com/@animadogi")
            webbrowser.open("https://www.tiktok.com/@animadogi")
            webbrowser.open("https://ko-fi.com/darkjake")
            webbrowser.open("https://github.com/Dark-Jake/TraductorTC")

        print("Reemplazo completado. El archivo modificado se guardó en:", archivo_salida)
        time.sleep(3)

reemplazar_palabras_tc()
